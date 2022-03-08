
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

dados = {'Compressive': '38.4',       
        'Cube Ref': '1A',
        'Date of Test': '09/02/2022',
        'Density': '2340',
        'Teste Age': '9'
        }

arquivo_excel = Workbook()
arquivo_excel = load_workbook('cube.xlsx')
excel = arquivo_excel.active
i=0
max_linha = excel.max_row
max_coluna = excel.max_column
excel_full = []
for linha in range(3, max_linha + 1):
    for column in range(1, max_coluna + 1):
        test =excel.cell(row=linha, column=column).value
        excel_full.append(test)
        print(column)
        input()
    # if excel_full[i] == dados['Cube Ref']:
    #     valores= { 
    #             "":"t",
    #             "t":"t",
    #             "t":"t",
    #             "t":"t",
    #             "t":"t"
    #     }
    #     i = i+10
    #     input()
    # else:
    #     print("CERTO")
    #     input()   





# if ws['A3'] == dados['Cliente Ref']:

#     print(ws['A3'].value)
#     print(dados['Cliente Ref'])
# else:
#     print('Error')